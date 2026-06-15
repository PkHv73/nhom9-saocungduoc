"""
Nhận JSON tickets từ stdin, xuất file Excel có định dạng chuẩn.
Usage: echo '{"tickets":[...]}' | python jira_export.py output.xlsx
"""
import json, sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def build_excel(data: dict, out_path: str):
    results = data.get("tickets", [])
    wb = Workbook()

    # ── Sheet 1: Chi tiết ──────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Chi tiết Tickets"
    ws.freeze_panes = "A2"

    HEADERS = ["Issue Key","Loại vấn đề","Root Cause","SLA Status","Ưu tiên","Escalate đến","Gợi ý xử lý CS"]
    COL_W   = [14,         26,           28,          11,          9,        28,             65]

    thin = Border(
        left=Side(style="thin", color="E2E8F0"), right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),  bottom=Side(style="thin", color="E2E8F0"),
    )
    hdr_fill  = PatternFill("solid", start_color="1E3A5F")
    hdr_font  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for ci, (h, w) in enumerate(zip(HEADERS, COL_W), 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = hdr_font; c.fill = hdr_fill; c.alignment = hdr_align; c.border = thin
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 30

    fill_breach = PatternFill("solid", start_color="FFF0F0")
    fill_infra  = PatternFill("solid", start_color="FFF8E7")
    fill_even   = PatternFill("solid", start_color="F0F4FF")
    fill_odd    = PatternFill("solid", start_color="FFFFFF")

    for i, r in enumerate(results, 2):
        is_breach = str(r.get("sla_status","")).upper() == "BREACH"
        is_infra  = "Infrastructure" in str(r.get("root_cause", r.get("type","")))
        is_prio   = r.get("priority_action", False)
        row_fill  = fill_breach if is_breach else (fill_infra if is_infra else (fill_even if i%2==0 else fill_odd))

        row_data = [
            r.get("key",""),
            r.get("type",""),
            r.get("root_cause", ""),
            r.get("sla_status","OK"),
            "🔥 Có" if is_prio else "Không",
            r.get("escalate_to",""),
            r.get("suggestion", r.get("error","")),
        ]
        for ci, val in enumerate(row_data, 1):
            c = ws.cell(row=i, column=ci, value=val)
            c.fill = row_fill; c.border = thin
            c.alignment = Alignment(vertical="top", wrap_text=(ci in (2,3,6,7)))
            if ci == 1:
                c.font = Font(name="Arial", size=9, bold=True, color="1E3A5F")
            elif ci == 4:
                c.font = Font(name="Arial", size=9, bold=True, color="DC2626" if is_breach else "166534")
            elif ci == 5:
                c.font = Font(name="Arial", size=9, bold=is_prio, color="C05621" if is_prio else "374151")
            else:
                c.font = Font(name="Arial", size=9)
        ws.row_dimensions[i].height = 45

    # ── Sheet 2: Tổng kết ─────────────────────────────────────────────────
    ws2 = wb.create_sheet("Tổng kết")
    ws2.column_dimensions["A"].width = 38
    ws2.column_dimensions["B"].width = 16

    from collections import Counter
    breach_n  = sum(1 for r in results if str(r.get("sla_status","")).upper()=="BREACH")
    prio_n    = sum(1 for r in results if r.get("priority_action"))
    rc_count  = Counter(r.get("root_cause","") for r in results)
    esc_count = Counter(r.get("escalate_to","") for r in results if r.get("escalate_to"))

    def hdr(row, col, val):
        c = ws2.cell(row=row, column=col, value=val)
        c.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", start_color="1E3A5F")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin

    def kv(row, k, v):
        ck = ws2.cell(row=row, column=1, value=k)
        ck.font = Font(name="Arial", bold=True, size=10)
        ck.fill = PatternFill("solid", start_color="EFF6FF")
        ck.border = thin; ck.alignment = Alignment(vertical="center")
        cv = ws2.cell(row=row, column=2, value=v)
        cv.font = Font(name="Arial", size=10, bold=True)
        cv.border = thin; cv.alignment = Alignment(horizontal="center", vertical="center")
        ws2.row_dimensions[row].height = 20

    hdr(1,1,"Chỉ số"); hdr(1,2,"Giá trị"); ws2.row_dimensions[1].height = 24
    kv(2,"Tổng số tickets", len(results))
    kv(3,"SLA Breach", breach_n)
    kv(4,"Cần ưu tiên xử lý", prio_n)
    kv(5,"Tỷ lệ SLA Breach", f"{breach_n/len(results)*100:.0f}%" if results else "0%")

    r = 7
    ws2.cell(row=r, column=1, value="Phân bố Root Cause").font = Font(name="Arial", bold=True, size=11, color="1E3A5F")
    hdr(r+1,1,"Root Cause"); hdr(r+1,2,"Số lượng"); ws2.row_dimensions[r+1].height = 22
    for ri2, (k2, v2) in enumerate(rc_count.most_common(), r+2):
        kv(ri2, k2, v2)

    r2 = r + 2 + len(rc_count) + 1
    ws2.cell(row=r2, column=1, value="Phân bố Escalate").font = Font(name="Arial", bold=True, size=11, color="1E3A5F")
    hdr(r2+1,1,"Team Escalate"); hdr(r2+1,2,"Số tickets"); ws2.row_dimensions[r2+1].height = 22
    for ri3, (k3, v3) in enumerate(esc_count.most_common(), r2+2):
        kv(ri3, k3, v3)

    wb.save(out_path)


if __name__ == "__main__":
    out = sys.argv[1] if len(sys.argv) > 1 else "jira_export_output.xlsx"
    raw = sys.stdin.read()
    data = json.loads(raw)
    build_excel(data, out)
    print(f"saved:{out}")
